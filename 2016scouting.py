from openpyxl import *
import Tkinter as tkinter

workbook_save_name = "test.xlsx"
workbook_load_name = "blank.xlsx"

# =AVERAGE(high!C2, high!E2, high!G2, high!I2, high!K2, high!M2, high!O2, high!Q2, high!S2, high!U2)
# =AVERAGE(low!C2, low!E2, low!G2, low!I2, low!K2, low!M2, low!O2, low!Q2, low!S2, low!U2)

team_list = [1, 192, 2, 3, 4, 5, 6, 7, 8, 9, 10,
11, 12, 13, 14, 15, 16, 17, 18, 19, 20,
21, 22, 23, 24, 25, 26, 27, 28, 29, 30,
31, 31, 33, 34, 35, 36, 37, 38, 39, 40,
41, 42, 43, 44, 45, 46, 47, 48, 49, 50,
51, 52, 53]

def init_no_sheet(worksheet):
	# initializes a worksheet that stores the number of matches recorded for each team
        # also used as a storage of whether a team can cross with a ball, an unused data value
	for row in range(2, len(team_list) + 2):
		worksheet.cell(row = row, column = 1, value = team_list[row - 2])
		if worksheet.cell(row = row, column = 2).value == None:
			worksheet.cell(row = row, column = 2).value = 0
	
	worksheet.cell(row = 1, column = 2, value = "Matches Played")
	worksheet.cell(row = 1, column = 3, value = "Cross w/ Ball?")

def init_shot_sheet(worksheet):
	# initalizes a worksheet that stores shot data (high or low)
        # contains column for attempts and successes for each match
	match = 1
	is_attempts = True
	for column in range(2, 22):
		if is_attempts:
			worksheet.cell(row = 1, column = column, value = "Match " + str(match) + " attempts")
		else:
			worksheet.cell(row = 1, column = column, value = "Match " + str(match) + " successes")
			match += 1
		is_attempts = not is_attempts

	for row in range(2, len(team_list) + 2):
		worksheet.cell(row = row, column = 1, value = team_list[row - 2])

def init_shot_analysis_sheet(worksheet):
	# intializes a worksheet that pulls shot data into excel functions for later reading
	def fill_single_function_column(column, sheet_function, range_start, range_end):
		# fills a column from row[range_start, range_end] with a certain sheet function
		for row in range(3, len(team_list) + 3):
			worksheet.cell(row = row, column = column, value = "=" + sheet_function + "(" + range_start + str(row) + ":" + range_end + str(row) + ")")

	Alphabet = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K",
	"L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
	
	worksheet.cell(row = 1, column = 2, value = "High Goal")
	worksheet.cell(row = 1, column = 17, value = "Low Goal")
	
	for row in range(3, len(team_list) + 3):
		worksheet.cell(row = row, column = 1, value = team_list[row - 3])
	
	for column in range(2, 12):
		worksheet.cell(row = 2, column = column, value = "Match " + str(column - 1) + " Average")
		attempts_index = (2 * (column - 2) + 1)
		attempts_letter = Alphabet[attempts_index]
		successes_letter = Alphabet[attempts_index + 1]
		for row in range(3, len(team_list) + 3):
			worksheet.cell(row = row, column = column, value = "=high!" + successes_letter + str(row - 1) + "/high!" + attempts_letter + str(row - 1))
	
	for column in range(17, 27):
		worksheet.cell(row = 2, column = column, value = "Match " + str(column - 1) + " Average")
		attempts_index = (2 * (column - 17) + 1)
		attempts_letter = Alphabet[attempts_index]
		successes_letter = Alphabet[attempts_index + 1]
		for row in range(3, len(team_list) + 3):
			worksheet.cell(row = row, column = column, value = "=low!" + successes_letter + str(row - 1) + "/low!" + attempts_letter + str(row - 1))
	
	worksheet.cell(row = 2, column = 12, value = "Average")
	worksheet.cell(row = 2, column = 13, value = "StDev")
	worksheet.cell(row = 2, column = 14, value = "Avg goals/match")
	worksheet.cell(row = 2, column = 15, value = "StDev goals/match")
	
	worksheet.cell(row = 2, column = 27, value = "Average")
	worksheet.cell(row = 2, column = 28, value = "StDev")
	worksheet.cell(row = 2, column = 29, value = "Avg goals/match")
	worksheet.cell(row = 2, column = 30, value = "StDev goals/match")
	
	#=AVERAGE(High!C2, High!E2, High!G2, High!I2, High!K2, High!M2, High!O2, High!Q2, High!S2, High!U2)

	fill_single_function_column(12, "AVERAGE", "B", "K")
	fill_single_function_column(13, "STDEV", "B", "K")
	fill_single_function_column(27, "AVERAGE", "Q", "Z")
	fill_single_function_column(28, "STDEV", "Z", "Q")	
	
def init_general_sheet(worksheet):
	# initializes a sheet for general data recording
	# team numbers are down the first column, match numbers are across the top row
	# includes avg and stdev functions for data analysis
	Alphabet = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K",
	"L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
	
	worksheet.cell(row = 1, column = len(team_list) + 2, value = "average")
	worksheet.cell(row = 1, column = len(team_list) + 3, value = "stdev")
	
	match = 1
	for column in range(2, 12):
		worksheet.cell(row = 1, column = column, value = "Match " + str(match))
		match += 1

	worksheet.cell(row = 1, column = 12, value = "avg")
	worksheet.cell(row = 1, column = 13, value = "stdev")
	
	for row in range(2, len(team_list) + 2):
		worksheet.cell(row = row, column = 1, value = team_list[row - 2])
		worksheet.cell(row = row, column = 12, value = "=AVERAGE(B" + str(row) + ":K" + str(row) + ")")
		worksheet.cell(row = row, column = 13, value = "=STDEV(B" + str(row) + ":K" + str(row) + ")")

def fill_shot_sheet(worksheet, team_number, matches_played, goal_values):
	# takes data from a match and fills it into a shot sheet created using init_shot_sheet
	team_index = team_list.index(team_number)

	entry_column = (matches_played * 2) + 2
	entry_row = team_index + 2

	worksheet.cell(row = entry_row, column = entry_column, value = int(goal_values[0]))
	worksheet.cell(row = entry_row, column = entry_column + 1, value = int(goal_values[1]))

def fill_defense_sheet(worksheet, team_number, matches_played, defense_crosses, defense_crosses_index):
	# takes data from a match and fills it into a defense sheet created using init_general_sheet
	team_index = team_list.index(team_number)

	entry_column = matches_played + 2
	entry_row = team_index + 2

	worksheet.cell(row = entry_row, column = entry_column, value = int(defense_crosses[defense_crosses_index]))

def fill_ball_cross(team_number, cross_ball_value):
	# takes data from a match about a team's ability to cross with a ball and fills it into the no sheet
        # not used in final iteration of scouting system
	entry_row = team_list.index(team_number) + 2

	if ((no_sheet.cell(row = entry_row, column = 3).value == None or
		no_sheet.cell(row = entry_row, column = 3).value == 0) and
		(cross_ball_value == 1)):
		no_sheet.cell(row = entry_row, column = 3).value = int(cross_ball_value)

def fill_auto_sheet(worksheet, team_number, matches_played, auto_values):
	# takes auton data from a match and fills it into a sheet created using init_general_sheet
        # 0 = nothing, 1 = reach, 2 = cross, 3 = cross and shoot low, 4 = cross and shoot high, 5 = 3 + recross, 6 = 4 + recross
	team_index = team_list.index(team_number)
	entry_column = matches_played + 2
	entry_row = team_index + 2
	#auto_choices = [reach_var, cross_var, low_var, high_var, recross_var, none_var]
	entry_value = 0

	entry = True
	while entry:
		if auto_values[5] == 1:
			entry_value = 0
		else:
			if auto_values[0] == 1:
				entry_value = 1
			if auto_values[1] == 1:
				entry_value = 2
			if auto_values[2] == 1:
				entry_value = 3
				if auto_values[4] == 1:
					entry_value = 5
					break
			if auto_values[3] == 1:
				entry_value = 4
				if auto_values[4] == 1:
					entry_value = 6
					break
		entry = False

	worksheet.cell(row = entry_row, column = entry_column, value = entry_value)

def fill_climb_sheet(worksheet, team_number, matches_played, climb_value):
	# takes climb capability data from a match and fills it into a sheet created with init_general_sheet
        # 0 = didn't attempt, 1 = attempt and fail, 2 = success
	team_index = team_list.index(team_number)
	entry_column = matches_played + 2
	entry_row = team_index + 2

	worksheet.cell(row = entry_row, column = entry_column, value = int(climb_value))

def fill_rip_sheet(worksheet, team_number, matches_played, rip_value):
	# takes data on whether a team died/lost comm from a match and fills it into a sheet created with init_general_sheet
        # 0 = didn't die, 1 = died
	team_index = team_list.index(team_number)
	entry_column = matches_played + 2
	entry_row = team_index + 2

	worksheet.cell(row = entry_row, column = entry_column, value = int(rip_value))

def data_entry(general_values, auto_values, shooting_values, defenses_chosen, defense_crosses, other_values):
	# takes full set of parsed match data from button_entry and calls data filling functions
	team_number = general_values[0]
	team_number = int(team_number)
	team_index = team_list.index(team_number)
	matches_played = no_sheet.cell(row = team_index + 2, column = 2).value
	
	matches_played = int(matches_played)

	high_values = [shooting_values[0], shooting_values[1]]
	low_values = [shooting_values[2], shooting_values[3]]
	print matches_played
	fill_shot_sheet(high_sheet, team_number, matches_played, high_values)
	fill_shot_sheet(low_sheet, team_number, matches_played, low_values)
	fill_auto_sheet(auton_sheet, team_number, matches_played, auto_values)
	fill_climb_sheet(climb_sheet, team_number, matches_played, other_values[1])
	fill_rip_sheet(rip_sheet, team_number, matches_played, other_values[2])
	fill_ball_cross(team_number, other_values[0])
	for i in range(0, len(defenses_chosen)):
		defense = defenses_chosen[i]
		fill_defense_sheet(scouting_data.get_sheet_by_name(defense), team_number, matches_played, defense_crosses, i)
	
	fill_defense_sheet(lowbar_sheet, team_number, matches_played, defense_crosses, len(defense_crosses) - 1)
	no_sheet.cell(row = team_index + 2, column = 2).value += 1
	scouting_data.save(workbook_save_name)

def button_entry(window, general_entry, auto_choices, shooting_entries, cat_choices, cat_entries, other_choices):
	# called by the enter button on the GUI
	# parses all entered values into a simplified format for data_entry
	
	#general_entry = [team_number_entry, match_number_entry]
	#auto_choices = [reach_var, cross_var, low_var, high_var, recross_var, none_var]
	#shooting_entries = [high_attempts_entry, high_successes_entry, low_attempts_entry, low_successes_entry]
	#cat_choices = [cat_a_choice, cat_b_choice, cat_c_choice, cat_d_choice]
	#cat_entries = [cat_a_entry, cat_b_entry, cat_c_entry, cat_d_entry, cat_e_entry]
	#other_choices = [cross_ball_var, climbing_choice, robot_RIP_var]
	general_values = []
	auto_values = []
	shooting_values = []
	defenses_chosen = []
	defense_crosses = []
	other_values = []
	data_entered = [general_values, auto_values, shooting_values,
		defenses_chosen, defense_crosses, other_values]
	for entry in general_entry:
		general_values.append(entry.get())
		entry.delete(0, tkinter.END)

	for var in auto_choices:
		auto_values.append(var.get())

	for entry in shooting_entries:
		shooting_values.append(entry.get())
		entry.delete(0, tkinter.END)

	for var in cat_choices:
		defenses_chosen.append(var.get())

	for entry in cat_entries:
		defense_crosses.append(entry.get())
		entry.delete(0, tkinter.END)

	for var in other_choices:
		other_values.append(var.get())

	data_entry(general_values, auto_values, shooting_values, 
			defenses_chosen, defense_crosses, other_values)

 	scouting_data.save(workbook_save_name)

def gui_init():
	# creates the GUI for entry
		
	window = tkinter.Tk()
	window.title(string = "GRT 2016 Scouting Data Input")
	window.geometry("1100x350")

	
	blank_label0 = tkinter.Label(window, text = "")
	blank_label1 = tkinter.Label(window, text = "")
	blank_label2 = tkinter.Label(window, text = "")
	blank_label3 = tkinter.Label(window, text = "")


	team_label = tkinter.Label(window, text = "Team Number")
	team_number_entry = tkinter.Entry(window)
	team_number_entry.insert(0, "192")

	match_label = tkinter.Label(window, text = "Match Number")
	match_number_entry = tkinter.Entry(window)

	general_entry = [team_number_entry, match_number_entry]
	
	high_attempts_entry = tkinter.Entry(window)
	high_successes_entry = tkinter.Entry(window)
	low_attempts_entry = tkinter.Entry(window)
	low_successes_entry = tkinter.Entry(window)
	
	cat_a_entry = tkinter.Entry(window)
	cat_b_entry = tkinter.Entry(window)
	cat_c_entry = tkinter.Entry(window)
	cat_d_entry = tkinter.Entry(window)
	cat_e_entry = tkinter.Entry(window)
		
	
	reach_var = tkinter.IntVar()
	cross_var = tkinter.IntVar()
	low_var = tkinter.IntVar()
	high_var = tkinter.IntVar()
	recross_var = tkinter.IntVar()
	none_var = tkinter.IntVar()
	auto_label = tkinter.Label(window, text = "Autonomous")
	auto_reach = tkinter.Checkbutton(window, text = "Reach", variable = reach_var)
	auto_cross = tkinter.Checkbutton(window, text = "Cross", variable = cross_var)
	auto_low = tkinter.Checkbutton(window, text = "Shoot low", variable = low_var)
	auto_high = tkinter.Checkbutton(window, text = "Shoot high", variable = high_var)
	auto_recross = tkinter.Checkbutton(window, text = "Recross", variable = recross_var)
	auto_none = tkinter.Checkbutton(window, text = "Nothing", variable = none_var)
	auto_choices = [reach_var, cross_var, low_var, high_var, recross_var, none_var]

	shooting_label = tkinter.Label(window, text = "Shooting")
	high_label = tkinter.Label(window, text = "High Attempts")
	low_label = tkinter.Label(window, text = "Low Attempts")
	attempts_label0 = tkinter.Label(window, text = "Attempts")
	successes_label0 = tkinter.Label(window, text = "Successes")
	attempts_label1 = tkinter.Label(window, text = "Attempts")
	successes_label1 = tkinter.Label(window, text = "Successes")
	
	shooting_entries = [high_attempts_entry, high_successes_entry,
		low_attempts_entry, low_successes_entry]
		
	for entry in shooting_entries:
		entry.insert(0, "0")
		
	
	
	cat_a_label = tkinter.Label(window, text = "Category A:")
	cat_a_choice = tkinter.Variable()
	portcullis_choice = tkinter.Radiobutton(window, text = "Portcullis", variable = cat_a_choice, value = "portcullis")
	cheval_choice = tkinter.Radiobutton(window, text = "Cheval de Frise", variable = cat_a_choice, value = "cheval de frise")

	cat_b_label = tkinter.Label(window, text = "Category B:")
	cat_b_choice = tkinter.Variable()
	moat_choice = tkinter.Radiobutton(window, text = "Moat", variable = cat_b_choice, value = "moat")
	ramparts_choice = tkinter.Radiobutton(window, text = "Ramparts", variable = cat_b_choice, value = "ramparts")
	
	cat_c_label = tkinter.Label(window, text = "Category C:")
	cat_c_choice = tkinter.Variable()
	drawbridge_choice = tkinter.Radiobutton(window, text = "Drawbridge", variable = cat_c_choice, value = "drawbridge")
	sally_port_choice = tkinter.Radiobutton(window, text = "Sally Port", variable = cat_c_choice, value = "sally port")
	
	cat_d_label = tkinter.Label(window, text = "Category D:")
	cat_d_choice = tkinter.Variable()
	rock_wall_choice = tkinter.Radiobutton(window, text = "Rock Wall", variable = cat_d_choice, value = "rock wall")
	rough_terrain_choice = tkinter.Radiobutton(window, text = "Rough Terrain", variable = cat_d_choice, value = "rough terrain")
	
	cat_e_label = tkinter.Label(window, text = "Category E:")
	low_bar_label = tkinter.Label(window, text = "Low Bar")
	

	cat_choices = [cat_a_choice, cat_b_choice,
		cat_c_choice, cat_d_choice]
	cat_entries = [cat_a_entry, cat_b_entry, cat_c_entry,
		cat_d_entry, cat_e_entry]
	
	for entry in cat_entries:
		entry.insert(0, "0")

	cross_ball_var = tkinter.Variable()
	cross_ball_check = tkinter.Checkbutton(window, text = "Cross w/ Ball?")

	climbing_label = tkinter.Label(window, text = "Climbing")
	climbing_choice = tkinter.Variable()
	climb_no_attempt = tkinter.Radiobutton(window, text = "Didn't attempt", variable = climbing_choice, value = 0)
	climb_attempt = tkinter.Radiobutton(window, text = "Attempt and fail", variable = climbing_choice, value = 1)
	climb_success = tkinter.Radiobutton(window, text = "Success", variable = climbing_choice, value = 2)

	robot_RIP_var = tkinter.IntVar()
	robot_RIP_button = tkinter.Checkbutton(window, text = "Did the robot break/lose comm?", variable = robot_RIP_var)

	other_choices = [cross_ball_var, climbing_choice, robot_RIP_var]

	button = tkinter.Button(window, text = "Enter", command = (lambda: button_entry(window, general_entry, auto_choices, shooting_entries, cat_choices, cat_entries, other_choices)))
	quit_button = tkinter.Button(window, text = "Quit", command = (lambda: quit_gui(window)))
	
	team_label.grid(row = 0, column = 0)
	team_number_entry.grid(row = 1, column = 0)

	match_label.grid(row = 0, column = 1)
	match_number_entry.grid(row = 1, column = 1)
	
	high_attempts_entry.grid(row = 1, column = 2)
	high_successes_entry.grid(row = 1, column = 3)

	low_attempts_entry.grid(row = 1, column = 4)
	low_successes_entry.grid(row = 1, column = 5)


	high_label.grid(row = 0, column = 2)
	successes_label0.grid(row = 0, column = 3)
	

	low_label.grid(row = 0, column = 4)
	successes_label1.grid(row = 0, column = 5)


	cat_a_label.grid(row = 3, column = 0)
	cat_a_entry.grid(row = 4, column = 0)
	portcullis_choice.grid(row = 5, column = 0)
	cheval_choice.grid(row = 6, column = 0)
	
	cat_b_label.grid(row = 3, column = 1)
	cat_b_entry.grid(row = 4, column = 1)
	moat_choice.grid(row = 5, column = 1)
	ramparts_choice.grid(row = 6, column = 1)

	cat_c_label.grid(row = 3, column = 2)
	cat_c_entry.grid(row = 4, column = 2)
	drawbridge_choice.grid(row = 5, column = 2)
	sally_port_choice.grid(row = 6, column = 2)

	cat_d_label.grid(row = 3, column = 3)
	cat_d_entry.grid(row = 4, column = 3)
	rock_wall_choice.grid(row = 5, column = 3)
	rough_terrain_choice.grid(row = 6, column = 3)

	cat_e_label.grid(row = 3, column = 4)
	cat_e_entry.grid(row = 4, column = 4)
	low_bar_label.grid(row = 5, column = 4)


	cross_ball_check.grid(row = 9, column = 0)
	robot_RIP_button.grid(row = 9, column = 2)

	climbing_label.grid(row = 8, column = 0)
	climb_no_attempt.grid(row = 8, column = 1)
	climb_attempt.grid(row = 8, column = 2)
	climb_success.grid(row = 8, column = 3)
	
	blank_label0.grid(row = 2, column = 0)
	blank_label1.grid(row = 7, column = 0)
	
	auto_label.grid(row = 11, column = 0)
	auto_reach.grid(row = 11, column = 1)
	auto_cross.grid(row = 11, column = 2)
	auto_low.grid(row = 11, column = 3)
	auto_high.grid(row = 12, column = 1)
	auto_recross.grid(row = 12, column = 2)
	auto_none.grid(row = 12, column = 3)
	
	button.grid(row = 0, column = 7)
	quit_button.grid(row = 1, column = 8)
	window.mainloop()
	tkinter.mainloop()

def quit_gui(window):
	scouting_data.save(workbook_save_name)
	window.quit()

scouting_data = load_workbook(workbook_load_name)

defense_sheets = [None, None, None, None, None, None, None, None, None]
# [portcullis, cheval, moat, ramparts, drawbridge, sally, workwall, roughterrain, lowbar]

if len(scouting_data.get_sheet_names()) < 3:
	# if loading from a blank or mostly blank workbook, create and assign all appropriate worksheets
	shot_analysis_sheet = scouting_data.create_sheet(title = "shot analysis")
	no_sheet = scouting_data.create_sheet(title = "no")
	init_no_sheet(no_sheet)
	
	auton_sheet = scouting_data.create_sheet(title = "auton")
	
	portcullis_sheet = scouting_data.create_sheet(title = "portcullis")
	cheval_sheet = scouting_data.create_sheet(title = "cheval de frise")
	moat_sheet = scouting_data.create_sheet(title = "moat")
	ramparts_sheet = scouting_data.create_sheet(title = "ramparts")
	drawbrdige_sheet = scouting_data.create_sheet(title = "drawbridge")
	sally_sheet = scouting_data.create_sheet(title = "sally port")
	rockwall_sheet = scouting_data.create_sheet(title = "rock wall")
	roughterrain_sheet = scouting_data.create_sheet(title = "rough terrain")
	lowbar_sheet = scouting_data.create_sheet(title = "lowbar")

	high_sheet = scouting_data.create_sheet(title = "high")
	low_sheet = scouting_data.create_sheet(title = "low")

	climb_sheet = scouting_data.create_sheet(title = "climb")
	rip_sheet = scouting_data.create_sheet(title = "rip")

	defense_sheets = [portcullis_sheet, cheval_sheet, moat_sheet,
		ramparts_sheet, drawbrdige_sheet, sally_sheet, rockwall_sheet,
		roughterrain_sheet, lowbar_sheet]

	init_general_sheet(auton_sheet)
	init_general_sheet(climb_sheet)
	init_general_sheet(rip_sheet)
	init_shot_sheet(high_sheet)
	init_shot_sheet(low_sheet)
	
	init_shot_analysis_sheet(shot_analysis_sheet)

	for defense_sheet in defense_sheets:
		init_general_sheet(defense_sheet)
else:
	# else, the workbook that is being loaded from is presumed to contain all appropriate sheets
	# the sheets are then assigned to appropriate variables
	for sheet in scouting_data.worksheets:
		if sheet.title == "portcullis":
			portcullis_sheet = sheet
			defense_sheets[0] = portcullis_sheet
		elif sheet.title == "cheval de frise":
			cheval_sheet = sheet
			defense_sheets[1] = cheval_sheet
		elif sheet.title == "moat":
			moat_sheet = sheet
			defense_sheets[2] = moat_sheet
		elif sheet.title == "ramparts":
			ramparts_sheet = sheet
			defense_sheets[3] = ramparts_sheet
		elif sheet.title == "drawbridge":
			drawbridge_sheet = sheet
			defense_sheets[4] = drawbridge_sheet
		elif sheet.title == "sally port":
			sally_sheet = sheet
			defense_sheets[5] = sally_sheet
		elif sheet.title == "rock wall":
			rockwall_sheet = sheet
			defense_sheets[6] = rockwall_sheet
		elif sheet.title == "rough terrain":
			roughterrain_sheet = sheet
			defense_sheets[7] = roughterrain_sheet
		elif sheet.title == "lowbar":
			lowbar_sheet = sheet
			defense_sheets[8] = lowbar_sheet
		elif sheet.title == "no":
			no_sheet = sheet
		elif sheet.title == "auton":
			auton_sheet = sheet
		elif sheet.title == "high":
			high_sheet = sheet
		elif sheet.title == "low":
			low_sheet = sheet
		elif sheet.title == "climb":
			climb_sheet = sheet
		elif sheet.title == "rip":
			rip_sheet = sheet
		elif sheet.title == "shot analysis":
			shot_analysis_sheet = sheet

sheet_names = scouting_data.get_sheet_names()

print sheet_names
scouting_data.save(workbook_save_name)

gui_init()

scouting_data.save(workbook_save_name)
